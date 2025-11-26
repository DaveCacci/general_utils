# USED TO REPLACE THE CONTENT OF AN EXCEL SHEET IN A FILE1 WITH THE CONTENT OF ANOTHER EXCEL SHEET CONTAINED IN FILE2 THAT HAS THE SAME SHEET NAME
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import logging

def replace_sheet_content(file1, file2):
    '''
    Replace the content of sheets in file1 with the content of sheets in file2 if they share the same name.
    Parameters:
    - file1: str, path to the first Excel file (to be modified)
    - file2: str, path to the second Excel file (source of new content)
    Note: from 24.11.2025 this function saves a backup of file1 before modifying it.
    '''
    # Load both Excel files
    wb1 = load_workbook(filename=file1)
    wb2 = load_workbook(filename=file2)

    # file1 is going to have its sheets replaced by the sheets in file2 (if they share the same name)
    # saving a copy of file1 before modifying it
    backup_file1 = file1.replace('.xlsx', '_beforemod.xlsx')
    wb1.save(backup_file1)
    logging.info(f"Backup of {file1} saved as {backup_file1} before modification.")
    
    # Get sheet names
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    
    # Iterate through the sheets in the second file
    for sheet in sheets2:
        if sheet in sheets1:
            # Read the sheet from the second file
            df = pd.read_excel(file2, sheet_name=sheet)
            
            # Remove the sheet from the first file
            wb1.remove(wb1[sheet])
            
            # Create a new sheet in the first file with the same name
            wb1.create_sheet(title=sheet)
            
            # Write the DataFrame to the new sheet in the first file
            for row in dataframe_to_rows(df, index=False, header=True):
                wb1[sheet].append(row)
    
    # Save the modified first file
    wb1.save(file1)
    logging.info(f"Content of {file1} replaced with content of {file2}")